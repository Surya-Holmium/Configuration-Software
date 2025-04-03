from PyQt6.QtWidgets import QApplication, QLabel, QWidget, QGridLayout, \
     QLineEdit, QPushButton, QComboBox, QMainWindow, QHBoxLayout,QMessageBox, \
     QWidgetAction, QFileDialog, QTextEdit, QToolBar, QStatusBar, QSizePolicy, \
     QGraphicsOpacityEffect, QCheckBox, QTableWidget, QTableWidgetItem, QDialog

import threading
import sys
from PyQt6.QtGui import QAction, QIcon, QPainter, QMovie, QPixmap, QBrush, QColor
from serial.tools.list_ports import comports
from PyQt6.QtCore import QTimer, QThread, pyqtSignal, QPropertyAnimation, Qt, QSettings, \
     QRect, Qt, QSize, pyqtSlot
import test
import subprocess
import time
import re
from openpyxl import Workbook, load_workbook
import datetime
import os

from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build

from enum import Enum

class STATE(Enum):
    CONNECTED = 1
    DISCONNECTED = 2
    TESTMODE = 3
    CONFIGUREMODE = 4
    CALIBRATEAI = 5
    CONFIGBUTTON = 6
    TESTRTC = 7
    TESTGSM = 8
    TESTWIFI = 9
    TESTETH = 10
    TESTSD = 11
    TESTMODRTU = 12
    TESTMODTCP = 13
    TESTALL = 14
    EXITNORMALLY = 15



currentState = STATE.DISCONNECTED.value


# If modifying these SCOPES, delete the file token.pickle.
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']

# The ID and range of your spreadsheet.
AUTHENTICATE_SPREADSHEET_ID = '1nBTVFEzVT6J5mbsFH935il6Byd60TCtUmU6Sx2vZjGc'
AUTHENTICATE_RANGE_NAME = 'Sheet1!A1:B'
DATABASE_SPREADSHEET_ID = '1ah5PPkq_2XMqDL99uXZdQ1oDzulAEWxfmiFCQ_ffsdw'
DATABASE_RANGE_NAME = 'Sheet1!A1:H'
DATABASE_RANGE_NAME1 = 'Sheet2!A1:B'
DATABASE_RANGE_NAME2 = 'Sheet3!A1:B'
DATABASE_RANGE_NAME3 = 'Sheet4!A1:B'
DATABASE_RANGE_NAME4 = 'Sheet5!A1:B'



class HandPointerMessageBox(QMessageBox):
    def showEvent(self, event):
        super().showEvent(event)
        for button in self.buttons():
            if isinstance(button, QPushButton):
                button.setCursor(Qt.CursorShape.PointingHandCursor)


class SerialThread(QThread):
    received = pyqtSignal(str)

    def __init__(self, port, baudrate):
        super().__init__()
        self.port = port
        self.baudrate = baudrate
        self.running = False

    def run(self):
        self.running = True
        self.mySerial = False
        try:
            self.ser = test.Serial(port=self.port, baudrate=self.baudrate)
            self.mySerial = True
            self.ser.dtr = False
            self.ser.rts = False

            while self.running:
                if self.running:
                    value = self.ser.readline()
                    try:
                        valueString = str(value.decode('UTF-8', errors='ignore'))
                        self.received.emit(valueString)
                    except UnicodeDecodeError as e:
                        print("Unicode error: ", e)
        except test.SerialException as e:
            print(f"Error connecting to {self.port}:{e}")
            self.received.emit(f"Error Connecting to {self.port}:{e}")
            
            

    def stop(self):
        self.running = False
        if self.mySerial:
            self.ser.close()

    def send_data(self, data):
        try:
            if self.running:
                if self.mySerial:
                    self.ser.write(data.encode(encoding="utf-8"))
        except Exception as e:
            print(f"Error sending data: {e}")


class UploadThread(QThread):
    output_received = pyqtSignal(str)

    def __init__(self, command):
        super().__init__()
        self.command = command

    def run(self):
        # Execute the command and capture the output
        process = subprocess.Popen(self.command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

        # Read and emit output in real-time
        for line in process.stdout:
            self.output_received.emit(line.strip())


class ImageLoader:
    def load_image(self, filename):
        try:
            filepath = sys._MEIPASS
        except Exception:
            filepath = os.path.abspath(".")
        return QPixmap(os.path.join(filepath, filename))
    
    def load_gif(self, filename):
        try:
            filepath = sys._MEIPASS
        except Exception:
            filepath = os.path.abspath(".")
        return QMovie(os.path.join(filepath, filename))
    

class GoogleSheetsAuthThread(QThread):
    auth_finished = pyqtSignal(object)

    def run(self):
        try:
            # Path to your service account key file
            if getattr(sys, 'frozen', False):  # Check if the app is running as a bundled executable
                service_account_file = os.path.join(sys._MEIPASS, 'credentialdata', 'htcudatabase-423608-d20a7aa9b6ad.json')
            else:
                service_account_file = os.path.join(os.path.dirname(__file__), 'credentialdata', 'htcudatabase-423608-d20a7aa9b6ad.json')

            creds = Credentials.from_service_account_file(service_account_file, scopes=SCOPES)
            service = build('sheets', 'v4', credentials=creds)
            self.auth_finished.emit(service)
        except Exception as e:
            print(f"Error during Google Sheets authentication: {e}")
            self.auth_finished.emit(None)
    

class LoginWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("HRMS Test & Config Utility")
        self.setMinimumSize(600, 400)
        self.setContentsMargins(200, 100, 200, 100)

        self.statusbar = self.statusBar()

        self.image_load = ImageLoader()

        self.window_icon = QIcon(self.image_load.load_image("icon\logo.png").scaled(60, 60))
        self.setWindowIcon(self.window_icon)

        # Create layout and add widgets
        layout = QGridLayout()
        
        # Create widgets
        self.message = QLabel("Welcome!")
        self.message.setStyleSheet("color: green; font-weight: bold; font-size: 16px")
        self.username_label = QLabel("Username:")
        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText("Enter Username")
        self.username_input.setStyleSheet("QLineEdit {background-color: white; border: 2px solid gray; border-radius: 10px; padding: 0 8px; }")

        self.password_label = QLabel("Password:")
        self.password_input = PasswordLineEdit()
        self.password_input.password_edit.setPlaceholderText("Enter Password")
        self.password_input.password_edit.setText(None)

        self.login_button = QPushButton("Login")
        self.login_button.setFixedSize(80, 35)
        self.login_button.setCursor(Qt.CursorShape.PointingHandCursor)
        self.login_button.setStyleSheet(
            """
            QPushButton {
                background-color: white;
                border: None;
                border-radius: 15px; 
                padding: 8px 16px;
                font-size: 12px;
            }

            QPushButton:hover {
                background-color: #A6F1F4; 
                border-color: grey; 
            }
            """
        )
        # self.login_button.setGeometry(200, 200, 200, 200)
        self.login_button.clicked.connect(self.handle_login)

        layout.addWidget(self.message, 0, 0, 1, 2, alignment=Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.username_label, 1, 0)
        layout.addWidget(self.username_input, 2, 0)
        layout.addWidget(self.password_label, 3, 0)
        layout.addWidget(self.password_input, 4, 0)
        layout.addWidget(self.login_button, 5, 0, 2, 1, alignment=Qt.AlignmentFlag.AlignCenter)

        # Create a central widget and set the layout
        central_widget = QWidget()
        central_widget.setLayout(layout)
        self.setCentralWidget(central_widget)

        self.service = None
        self.auth_thread = GoogleSheetsAuthThread()
        self.auth_thread.auth_finished.connect(self.on_auth_finished)
        self.auth_thread.start()

    def on_auth_finished(self, service):
        self.service = service

    def handle_login(self):
        if not self.service:
            QMessageBox.warning(self, "Error", "Authentication is still in progress. Please wait.")
            return
        
        username = self.username_input.text()
        password = self.password_input.password_edit.text()

        # Check if the user exists in the Google Sheet
        user_exists = self.check_user_in_sheet(username, password)
        if user_exists:
            # QMessageBox.information(self, "Login Successful", "You have successfully logged in.")
            self.open_new_window()
        else:
            QMessageBox.warning(self, "Login Failed", "The username or password is incorrect or if you are new user then please register")

    def check_user_in_sheet(self, username, password):
        try:
            sheet = self.service.spreadsheets()
            result = sheet.values().get(spreadsheetId=AUTHENTICATE_SPREADSHEET_ID, range=AUTHENTICATE_RANGE_NAME).execute()
            values = result.get('values', [])
            # print(values)
            for row in values:
                if len(row) >= 2 and row[0] == username and row[1] == password:
                    return True
            return False
        except Exception as e:
            print(f"Error accessing Google Sheets: {e}")
            return False

    def open_new_window(self):
        self.new_window = SerialMonitor(self.service)
        self.new_window.show()
        self.new_window.setStyleSheet("background-color: #add8e6;")
        self.close()


class  SerialMonitor(QMainWindow):
    def __init__(self, service):
        super().__init__()
        self.service = service
        try:
            self.setWindowTitle("HRMS Test & Config Utility")
            self.setMinimumSize(600, 400)

            self.statusbar = self.statusBar()

            self.terminalWindow = None  # Initialize terminal window reference
            self.configWindow = None
            self.testWindow = None
            self.calibrateAIWindow = None
            self.connection_open = False    
            self.informationwindow = None
            self.programWindow = None

            self.comboBox = QComboBox()
            self.comboBox.setCursor(Qt.CursorShape.PointingHandCursor)
            self.comboBox.setStyleSheet("background-color: white")

            self.comboBox.setPlaceholderText("Select COM Port...")

            self.addComboBoxToMenuBar()

            self.selected_port = None

            self.timer = QTimer(self)
            self.timer.timeout.connect(self.scan_USBPort)
            self.timer.start(1000)

            self.baudrate = QComboBox()
            self.baudrate.setCursor(Qt.CursorShape.PointingHandCursor)
            self.baudrate.setStyleSheet("background-color: white;")
            baudrates = ["115200", "9600"]
            self.baudrate.addItems(baudrates)

            self.addBaudrateToMenuBar()

            self.connect_button_menu = self.menuBar().addMenu('&Connection')
            self.connect_button_menu.setCursor(Qt.CursorShape.PointingHandCursor)
            self.connect_button_menu.setStyleSheet("background-color: white; color: black;")

            self.connect_button = QAction("Connect", self)
            self.connect_button_menu.addAction(self.connect_button)
            self.connect_button.triggered.connect(self.on_connect_clicked)
            self.connect_button.hovered.connect(lambda: self.setCursor(Qt.CursorShape.PointingHandCursor))
            self.connect_button.hovered.connect(lambda: QApplication.restoreOverrideCursor())

            self.disconnect_button = QAction( "Disconnect", self)
            self.connect_button_menu.addAction(self.disconnect_button)
            self.disconnect_button.triggered.connect(self.on_disconnect_clicked)
            self.disconnect_button.hovered.connect(lambda: self.setCursor(Qt.CursorShape.PointingHandCursor))
            self.disconnect_button.hovered.connect(lambda: QApplication.restoreOverrideCursor())

            self.help_menu_item = self.menuBar().addMenu('&Help')
            self.help_menu_item.setCursor(Qt.CursorShape.PointingHandCursor)
            self.help_menu_item.setStyleSheet("color: green;")
                                                
            about_action = QAction("About", self)
            self.help_menu_item.addAction(about_action)

            #If help menu not showing
            # about_action.setMenuRole(QAction.MenuRole.NoRole)
            about_action.triggered.connect(self.about)
            about_action.hovered.connect(lambda: self.setCursor(Qt.CursorShape.PointingHandCursor))
            about_action.hovered.connect(lambda: QApplication.restoreOverrideCursor())


            self.program = QAction("Program", self)
            # program.setStatusTip("program")
            self.program.triggered.connect(self.programFW)
            self.program.hovered.connect(lambda: self.setCursor(Qt.CursorShape.PointingHandCursor))
            self.program.hovered.connect(lambda: QApplication.restoreOverrideCursor())


            self.config = QAction("Configure Device", self)
            self.config.setEnabled(False)  # Disable the action initially
            self.config.triggered.connect(self.configureDevice)
            self.config.hovered.connect(lambda: self.setCursor(Qt.CursorShape.PointingHandCursor))
            self.config.hovered.connect(lambda: QApplication.restoreOverrideCursor())

            self.test = QAction("Test Device", self)
            self.test.setEnabled(False)   # Disable the action initially
            self.test.triggered.connect(self.testDevice)
            self.test.hovered.connect(lambda: self.setCursor(Qt.CursorShape.PointingHandCursor))
            self.test.hovered.connect(lambda: QApplication.restoreOverrideCursor())

            self.calibrate_ai = QAction("Calibrate AI", self)
            self.calibrate_ai.setEnabled(False)   # Disable the action initially
            self.calibrate_ai.triggered.connect(self.calibrate_AI)
            self.calibrate_ai.hovered.connect(lambda: self.setCursor(Qt.CursorShape.PointingHandCursor))
            self.calibrate_ai.hovered.connect(lambda: QApplication.restoreOverrideCursor())

            self.exit = QAction("Exit", self)
            self.exit.setEnabled(False)   # Disable the action initially
            self.exit.triggered.connect(self.exit_All)
            self.exit.hovered.connect(lambda: self.setCursor(Qt.CursorShape.PointingHandCursor))
            self.exit.hovered.connect(lambda: QApplication.restoreOverrideCursor())

            self.show_data = QAction("Show Data", self)
            # self.show_data.setEnabled(False)   # Disable the action initially
            self.show_data.triggered.connect(self.show_DataWindow)
            self.show_data.hovered.connect(lambda: self.setCursor(Qt.CursorShape.PointingHandCursor))
            self.show_data.hovered.connect(lambda: QApplication.restoreOverrideCursor())

            self.image_load = ImageLoader()
            icon = QIcon(self.image_load.load_image("icon\magnifying-glass.png"))

            self.serialbutton = QAction(icon,"SerialButton", self)
            # self.serialbutton.setStyleSheet("QAction { border: 2px solid gray; border-radius: 10px; padding: 0 8px; }")
            self.serialbutton.triggered.connect(self.openTerminalWindow)
            self.serialbutton.hovered.connect(lambda: self.setCursor(Qt.CursorShape.PointingHandCursor))
            self.serialbutton.hovered.connect(lambda: QApplication.restoreOverrideCursor())

            #Creating toolbar and adding toolbar elements
            toolbar = QToolBar()
            toolbar.setStyleSheet("background-color: #D4F1F4;")
            toolbar.setMovable(True)
            self.addToolBar(toolbar)

            toolbar.addActions([self.program, self.config, self.test, self.calibrate_ai, self.exit, self.show_data])
            
            # Add a spacer item to push the serialbutton to the corner
            spacer = QWidget()
            spacer.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)


            toolbar.addWidget(spacer)
            toolbar.addAction(self.serialbutton)

            self.programbutton = QPushButton("Program")
            self.programbutton.clicked.connect(self.program.trigger)

            self.configbutton = QPushButton("Configure Device")
            self.configbutton.clicked.connect(self.config.trigger)

            self.testbutton = QPushButton("Test Device")
            self.testbutton.clicked.connect(self.test.trigger)

            self.calibratebutton = QPushButton("Calibrate AI")
            self.calibratebutton.clicked.connect(self.calibrate_ai.trigger)

            self.showdatabutton = QPushButton("Show Data")
            self.showdatabutton.clicked.connect(self.show_data.trigger)

            self.exitbutton = QPushButton("Exit")
            self.exitbutton.clicked.connect(self.exit.trigger)

            self.serialmonitorbutton = QPushButton("SerialButton")
            self.serialmonitorbutton.clicked.connect(self.serialbutton.trigger)

            # Apply styles
            self.apply_styles()

            self.selected_port = None
            self.serial_thread = None
            # self.terminalWindow = None

            #Creating status bar and adding status bar element
            self.statusbar = QStatusBar()
            self.statusbar.setStyleSheet("background-color: #D4F1F4; color: green;  font-weight: bold; font-size: 16px;")
            self.setStatusBar(self.statusbar)
            
            self.window_icon = QIcon(self.image_load.load_image("icon\logo.png").scaled(60, 60))
            self.setWindowIcon(self.window_icon)

        except AttributeError as e:
            # Handle the AttributeError appropriately
            print(f"AttributeError occurred: {e}")

    def apply_styles(self):
        # Style the buttons
        button_style = """
            QPushButton {
                background-color: white; 
                border: 2px solid gray; 
                border-radius: 10px; 
                padding: 0 8px; 
            }
            QPushButton:hover {
                background-color: #A6F1F4; 
            }
        """
        self.programbutton.setStyleSheet(button_style)
        self.configbutton.setStyleSheet(button_style)
        self.testbutton.setStyleSheet(button_style)
        self.calibratebutton.setStyleSheet(button_style)
        self.showdatabutton.setStyleSheet(button_style)
        self.exitbutton.setStyleSheet(button_style)
        self.serialmonitorbutton.setStyleSheet(button_style)

        # Style the toolbar buttons (QToolButton) associated with QActions
        toolbar_style = """
            QToolButton {
                background-color: white; 
                border: 2px solid gray; 
                border-radius: 10px; 
                padding: 0 8px; 
            }
            QToolButton:hover {
                background-color: #A6F1F4; 
            }
        """
        self.findChildren(QToolBar)[0].setStyleSheet(toolbar_style)

    def about(self):
        dialog = AboutDialog(self.image_load)
        dialog.exec()

    def addComboBoxToMenuBar(self):
        # Create a menu bar
        menu_bar = self.menuBar()

        # Create a menu
        menu = menu_bar.addMenu("Select Port")
        menu.setCursor(Qt.CursorShape.PointingHandCursor)

        # Add a custom widget containing the ComboBox to the menu
        combo_box_widget = QWidgetAction(self)
        combo_box_widget.setDefaultWidget(self.comboBox)
        menu.addAction(combo_box_widget)


    def addBaudrateToMenuBar(self):
        menu = self.menuBar().addMenu("Set Baudrate")
        menu.setCursor(Qt.CursorShape.PointingHandCursor)

        baudrate_widget = QWidgetAction(self)
        baudrate_widget.setDefaultWidget(self.baudrate)
        menu.addAction(baudrate_widget)

    def scan_USBPort(self):
        """Scan the USB ports and return a list of available port names."""

        current_port = self.comboBox.currentText()

        # Clear existing port names
        self.comboBox.clear()

        # Get list of available ports
        ports = comports()

        # Add port names to the combo box
        if not ports:
            # print("Ports not found")
            return 
        else:
            for port_info in ports:
                port_name = port_info.device
                self.comboBox.addItem(port_name)

        # Restore the previously selected port if it still exists
        if current_port:
            index = self.comboBox.findText(current_port)
            if index != -1:
                self.comboBox.setCurrentIndex(index)
            else:
                self.selected_port = None
        

    def on_connect_clicked(self):
        # Fetch the selected port from the combo box
        self.selected_port = self.comboBox.currentText()
        self.selected_baudrate = self.baudrate.currentText()
        # print("port = ",self.selected_port)
        global currentState
        currentState = STATE.CONNECTED.value
        # Check if a port is selected
        if not self.selected_port:
            msg_box = HandPointerMessageBox()
            msg_box.setWindowTitle("Warning")
            msg_box.setWindowIcon(self.window_icon)
            msg_box.setText("Please select a port.")
            msg_box.setIcon(QMessageBox.Icon.Warning)
            msg_box.exec()
            return
        
        # Create and start the serial thread
        self.serial_thread = SerialThread(self.selected_port, self.selected_baudrate)
        self.serial_thread.received.connect(self.on_data_received)  # Connect signal to data received slot   
        try:
            self.serial_thread.start()
            self.connection_open = True
        except AttributeError as e:
            msg_box = HandPointerMessageBox()
            msg_box.setWindowTitle("Warning")
            msg_box.setWindowIcon(self.window_icon)
            msg_box.setText("Error", f"Error connecting to the device: {str(e)}")
            msg_box.setIcon(QMessageBox.Icon.Warning)
            msg_box.exec()

        # self.serial_thread.received.connect(self.load_data_toScreen)

        threading.Thread(target=self.send_data_toGetMode).start()
        
    def send_data_toGetMode(self):
        time.sleep(4)
        self.serial_thread.send_data("HRMS-1810" + "\n")


    def on_data_received(self, data):
        # Append received data to the QTextEdit box of TerminalWindow
        global currentState
        self.data = data
        print(self.data)
        try:
            if self.terminalWindow is not None:
                if self.connection_open:
                    self.terminalWindow.serial_text.append(self.data)
        except AttributeError as e:
            print(f"Attribute Error in terminal window method call: {str(e)}")

        if currentState == STATE.CONNECTED.value:
            # self.informationwindow.table = None
            if "Holmium Technologies" in self.data[1:22]:
                self.informationwindow = InformationWindow()
                self.setCentralWidget(self.informationwindow)

            elif "serialNo" in self.data.split(":")[0]:
                self.update_table_item(0, "Serial No")

            elif "modelNo" in self.data.split(":")[0]:
                self.update_table_item(1, "Model No")

            elif "firmwareVersion" in self.data.split(":")[0]:
                self.update_table_item(2, "Firmware Version")

            elif "[Holmium Technologies Pvt. Ltd.]" in self.data:
                self.config.setEnabled(True)
                self.test.setEnabled(True)
                self.calibrate_ai.setEnabled(True)
                self.exit.setEnabled(True)
                self.statusbar.clearMessage()
                if self.connect_button.isChecked:
                    self.statusbar.showMessage('Connected')

            elif f"Error Connecting to {self.selected_port}" in self.data:
                msg_box = HandPointerMessageBox()
                msg_box.setWindowTitle("Warning")
                msg_box.setWindowIcon(self.window_icon)
                msg_box.setText(self.data)
                msg_box.setIcon(QMessageBox.Icon.Warning)
                msg_box.exec()

        elif  currentState == STATE.CONFIGUREMODE.value:
            if "Enter New Serial No" in self.data:
                self.test.setEnabled(False)
                self.calibrate_ai.setEnabled(False)
                self.exit.setEnabled(False)
                self.statusbar.setStyleSheet("background-color: #D4F1F4; color: green;  font-weight: bold; font-size: 18px;")

        elif currentState == STATE.CONFIGBUTTON.value:
            if "device data saved successfully; now resetting" in self.data:
                self.statusbar.showMessage("Device Configuration Successful")

        elif currentState == STATE.TESTMODE.value:
            if "-->Inside Test Mode" in self.data:
                self.config.setEnabled(False)
                self.calibrate_ai.setEnabled(False)
                self.exit.setEnabled(False)
                self.statusbar.setStyleSheet("background-color: #D4F1F4; color: green;  font-weight: bold; font-size: 18px;")
        
        elif currentState == STATE.TESTRTC.value:
            if ">>> Testing RTC <<<" in self.data:
                self.testWindow.testrtc.setStyleSheet("QPushButton {background-color:#EBE846; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;}")

            elif ">>> RTC Test OK <<<" in self.data:
                self.statusbar.clearMessage()
                self.testWindow.testrtc.setStyleSheet("background-color: #26D07C; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
                self.testWindow.movie_label.setVisible(False)
                currentState = STATE.TESTMODE.value
                    
            elif ">>> RTC Test Failed! <<<" in self.data:
                self.testWindow.testrtc.setStyleSheet("background-color : #FF7276; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
                # self.statusbar.setStyleSheet("background-color: #D4F1F4; font-weight: bold; font-size: 18px;")
                self.testWindow.movie_label.setVisible(False)
                currentState = STATE.TESTMODE.value
            else:
                self.statusbar.clearMessage()
                self.statusbar.showMessage(self.data)
        
        elif currentState == STATE.TESTGSM.value:
            if ">>> Testing GSM <<<" in self.data:
                self.testWindow.testgsm.setStyleSheet("QPushButton {background-color: #EBE846; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;}")

            elif ">>> GSM Test OK <<<" in self.data:
                self.statusbar.clearMessage()
                self.testWindow.testgsm.setStyleSheet("background-color: #26D07C; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
                self.testWindow.movie_label.setVisible(False)
                currentState = STATE.TESTMODE.value

            elif ">>> GSM Test Failed! <<<" in self.data:
                self.testWindow.testgsm.setStyleSheet("background-color : #FF7276; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
                # self.statusbar.setStyleSheet("background-color: #D4F1F4; font-weight: bold; font-size: 18px;")
                self.testWindow.movie_label.setVisible(False)
                currentState = STATE.TESTMODE.value
            else:
                self.statusbar.clearMessage()
                self.statusbar.showMessage(self.data)

        elif currentState == STATE.TESTWIFI.value:
            if ">>> Testing WiFi <<<" in self.data:
                self.testWindow.testwifi.setStyleSheet("QPushButton {background-color: #EBE846; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;}")
            
            elif ">>> WiFi Test OK <<<" in self.data:
                self.statusbar.clearMessage()
                self.testWindow.testwifi.setStyleSheet("background-color: #26D07C; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
                self.testWindow.movie_label.setVisible(False)
                currentState = STATE.TESTMODE.value
            elif ">>> WiFi Test Failed! <<<" in self.data:
                self.testWindow.testwifi.setStyleSheet("background-color : #FF7276; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
                # self.statusbar.setStyleSheet("background-color: #D4F1F4; font-weight: bold; font-size: 18px;")
                self.testWindow.movie_label.setVisible(False)
                currentState = STATE.TESTMODE.value
            else:
                self.statusbar.clearMessage()
                self.statusbar.showMessage(self.data)

        elif currentState == STATE.TESTETH.value:
            if ">>> Testing Ethernet <<<" in self.data:
                self.testWindow.testethernet.setStyleSheet("QPushButton {background-color: #EBE846; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;}")
            
            elif ">>> Ethernet Test OK <<<" in self.data:
                self.statusbar.clearMessage()
                self.testWindow.testethernet.setStyleSheet("background-color: #26D07C; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
                self.testWindow.movie_label.setVisible(False)
                currentState = STATE.TESTMODE.value
            elif ">>> Ethernet Test Failed! <<<" in self.data:
                self.testWindow.testethernet.setStyleSheet("background-color : #FF7276; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
                # self.statusbar.setStyleSheet("background-color: #D4F1F4; font-weight: bold; font-size: 18px;")
                self.testWindow.movie_label.setVisible(False)
                currentState = STATE.TESTMODE.value
            else:
                self.statusbar.clearMessage()
                self.statusbar.showMessage(self.data)

        elif currentState == STATE.TESTSD.value:
            if ">>> Testing SD Card <<<" in self.data:
                self.testWindow.testsdcard.setStyleSheet("QPushButton {background-color: #EBE846; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;}")
            
            elif ">>> SD Test OK <<<" in self.data:
                self.statusbar.clearMessage()
                self.testWindow.testsdcard.setStyleSheet("background-color: #26D07C; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
                self.testWindow.movie_label.setVisible(False)
                currentState = STATE.TESTMODE.value
            elif ">>> SD Test Failed! <<<" in self.data:
                self.testWindow.testsdcard.setStyleSheet("background-color : #FF7276; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
                # self.statusbar.setStyleSheet("background-color: #D4F1F4; font-weight: bold; font-size: 18px;")
                self.testWindow.movie_label.setVisible(False)
                currentState = STATE.TESTMODE.value
            else:
                self.statusbar.clearMessage()
                self.statusbar.showMessage(self.data)
            
        elif currentState == STATE.TESTMODRTU.value:
            if ">>> Testing Modbus RTU 1 <<<" in self.data:
                self.testWindow.testmodbusrtu.setStyleSheet("QPushButton {background-color: #EBE846; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;}")
            
            elif ">>> Modbus RTU 1 Test OK <<<" in self.data:
                self.statusbar.clearMessage()
                self.testWindow.testmodbusrtu.setStyleSheet("background-color: #26D07C; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
                # self.testWindow.movie_label.setVisible(False)
                # currentState = STATE.TESTMODE.value
            elif ">>> Modbus RTU 1 Test Failed! <<<" in self.data:
                self.testWindow.testmodbusrtu.setStyleSheet("background-color : #FF7276; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
                # self.statusbar.setStyleSheet("background-color: #D4F1F4; font-weight: bold; font-size: 18px;")
                # self.testWindow.movie_label.setVisible(False)
                # currentState = STATE.TESTMODE.value

            elif ">>> Testing Modbus RTU 2 <<<" in self.data:
                self.testWindow.testmodbusrtu.setStyleSheet("QPushButton {background-color: #EBE846; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;}")

            elif ">>> Modbus RTU 2 Test OK <<<" in self.data:
                self.statusbar.clearMessage()
                self.testWindow.testmodbusrtu.setStyleSheet("background-color: #26D07C; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
                self.testWindow.movie_label.setVisible(False)
                currentState = STATE.TESTMODE.value
            elif ">>> Modbus RTU 2 Test Failed! <<<" in self.data:
                self.testWindow.testmodbusrtu.setStyleSheet("background-color : #FF7276; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
                # self.statusbar.setStyleSheet("background-color: #D4F1F4; font-weight: bold; font-size: 18px;")
                self.testWindow.movie_label.setVisible(False)
                currentState = STATE.TESTMODE.value
            else:
                self.statusbar.clearMessage()
                self.statusbar.showMessage(self.data)

        elif currentState == STATE.TESTMODTCP.value:
            if ">>> Testing Modbus TCP <<<" in self.data:
                self.testWindow.testmodbustcp.setStyleSheet("QPushButton {background-color: #EBE846; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;}")

            elif ">>> Modbus TCP Test OK <<<" in self.data:
                self.statusbar.clearMessage()
                self.testWindow.testmodbustcp.setStyleSheet("background-color: #26D07C; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
                self.testWindow.movie_label.setVisible(False)
                currentState = STATE.TESTMODE.value
            elif ">>> Modbus TCP Test Failed! <<<" in self.data:
                self.testWindow.testmodbustcp.setStyleSheet("background-color : #FF7276; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
                # self.statusbar.setStyleSheet("background-color: #D4F1F4; font-weight: bold; font-size: 18px;")
                self.testWindow.movie_label.setVisible(False)
                currentState = STATE.TESTMODE.value
            else:
                self.statusbar.clearMessage()
                self.statusbar.showMessage(self.data)
                
        elif currentState == STATE.TESTALL.value:
            if ">>> Testing All <<<" in self.data:
                self.testWindow.testrtc.setStyleSheet("QPushButton {background-color:#FFFFFF; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;}")
                self.testWindow.testgsm.setStyleSheet("QPushButton {background-color:#FFFFFF; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;}")
                self.testWindow.testwifi.setStyleSheet("QPushButton {background-color:#FFFFFF; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;}")
                self.testWindow.testethernet.setStyleSheet("QPushButton {background-color:#FFFFFF; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;}")
                self.testWindow.testsdcard.setStyleSheet("QPushButton {background-color:#FFFFFF; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;}")
                self.testWindow.testmodbusrtu.setStyleSheet("QPushButton {background-color:#FFFFFF; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;}")
                self.testWindow.testmodbustcp.setStyleSheet("QPushButton {background-color:#FFFFFF; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;}")

            elif ">>> Testing RTC <<<" in self.data:
                self.testWindow.testrtc.setStyleSheet("QPushButton {background-color:#EBE846; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;}")
            elif ">>> RTC Test OK <<<" in self.data:
                self.testWindow.testrtc.setStyleSheet("background-color: #26D07C; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")                   
            elif ">>> RTC Test Failed! <<<" in self.data:
                self.testWindow.testrtc.setStyleSheet("background-color : #FF7276; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
                # self.statusbar.setStyleSheet("background-color: #D4F1F4; font-weight: bold; font-size: 18px;")

            elif ">>> Testing GSM <<<" in self.data:
                self.testWindow.testgsm.setStyleSheet("QPushButton {background-color:#EBE846; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;}")
            elif ">>> GSM Test OK <<<" in self.data:
                self.testWindow.testgsm.setStyleSheet("background-color: #26D07C; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
            elif ">>> GSM Test Failed! <<<" in self.data:
                self.testWindow.testgsm.setStyleSheet("background-color : #FF7276; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
                # self.statusbar.setStyleSheet("background-color: #D4F1F4; font-weight: bold; font-size: 18px;")

            elif ">>> Testing WiFi <<<" in self.data:
                self.testWindow.testwifi.setStyleSheet("QPushButton {background-color:#EBE846; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;}") 
            elif ">>> WiFi Test OK <<<" in self.data:
                self.testWindow.testwifi.setStyleSheet("background-color: #26D07C; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
            elif ">>> WiFi Test Failed! <<<" in self.data:
                self.testWindow.testwifi.setStyleSheet("background-color : #FF7276; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
                # self.statusbar.setStyleSheet("background-color: #D4F1F4;  font-weight: bold; font-size: 18px;")

            elif ">>> Testing Ethernet <<<" in self.data:
                self.testWindow.testethernet.setStyleSheet("QPushButton {background-color:#EBE846; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;}")
            elif ">>> Ethernet Test OK <<<" in self.data:
                self.testWindow.testethernet.setStyleSheet("background-color: #26D07C; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
            elif ">>> Ethernet Test Failed! <<<" in self.data:
                self.testWindow.testethernet.setStyleSheet("background-color : #FF7276; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
                # self.statusbar.setStyleSheet("background-color: #D4F1F4;  font-weight: bold; font-size: 18px;")

            elif ">>> Testing SD Card <<<" in self.data:
                self.testWindow.testsdcard.setStyleSheet("QPushButton {background-color:#EBE846; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;}")
            elif ">>> SD Test OK <<<" in self.data:
                self.testWindow.testsdcard.setStyleSheet("background-color: #26D07C; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
            elif ">>> SD Test Failed! <<<" in self.data:
                self.testWindow.testsdcard.setStyleSheet("background-color : #FF7276; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
                # self.statusbar.setStyleSheet("background-color: #D4F1F4; font-weight: bold; font-size: 18px;")

            elif ">>> Testing Modbus RTU 1 <<<" in self.data:
                self.testWindow.testmodbusrtu.setStyleSheet("QPushButton {background-color:#EBE846; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;}")
            elif ">>> Modbus RTU 1 Test OK <<<" in self.data:
                self.testWindow.testmodbusrtu.setStyleSheet("background-color: #26D07C; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
            elif ">>> Modbus RTU 1 Test Failed! <<<" in self.data:
                self.testWindow.testmodbusrtu.setStyleSheet("background-color : #FF7276; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
                # self.statusbar.setStyleSheet("background-color: #D4F1F4; font-weight: bold; font-size: 18px;")
            elif ">>> Testing Modbus RTU 2 <<<" in self.data:
                self.testWindow.testmodbusrtu.setStyleSheet("QPushButton {background-color:#EBE846; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;}")
            elif ">>> Modbus RTU 2 Test OK <<<" in self.data:
                self.testWindow.testmodbusrtu.setStyleSheet("background-color: #26D07C; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
            elif ">>> Modbus RTU 2 Test Failed! <<<" in self.data:
                self.testWindow.testmodbusrtu.setStyleSheet("background-color : #FF7276; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
                # self.statusbar.setStyleSheet("background-color: #D4F1F4; font-weight: bold; font-size: 18px;")

            elif ">>> Testing Modbus TCP <<<" in self.data:
                self.testWindow.testmodbustcp.setStyleSheet("QPushButton {background-color:#EBE846; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;}")
            elif ">>> Modbus TCP Test OK <<<" in self.data:
                self.testWindow.testmodbustcp.setStyleSheet("background-color: #26D07C; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
            elif ">>> Modbus TCP Test Failed! <<<" in self.data:
                self.testWindow.testmodbustcp.setStyleSheet("background-color : #FF7276; border: None; border-radius: 15px; padding: 8px 16px; font-size: 14px;")
                # self.statusbar.setStyleSheet("background-color: #D4F1F4; font-weight: bold; font-size: 18px;")
            
            elif ">>> All Tests Done <<<" in self.data:
                self.statusbar.showMessage("All Tests Done, Please, do the separete test for each test module if you find any error in the test Modules.")
                # self.statusbar.setStyleSheet("background-color: #D4F1F4; color: green;  font-weight: bold; font-size: 18px;")
                self.testWindow.movie_testAll.setVisible(False)
                currentState = STATE.TESTMODE.value
            else:
                self.statusbar.clearMessage()
                self.statusbar.showMessage(self.data)

        elif currentState == STATE.CALIBRATEAI.value:
            if "please wait" in self.data:
                self.config.setEnabled(False)
                self.test.setEnabled(False)
                self.exit.setEnabled(False)

        elif currentState == STATE.EXITNORMALLY.value:
            self.statusbar.showMessage(self.data) 

    def update_table_item(self, row, label):
        try:
            if self.informationwindow is not None and self.informationwindow.table is not None:
                col = 1
                if row < self.informationwindow.table.rowCount():
                    items = self.informationwindow.table.item(row, col)
                    if items is None:
                        value = QTableWidgetItem("   " + self.data.split(":")[1].strip())
                        self.informationwindow.table.setItem(row, col, value)
            else:
                print(f"Error: InformationWindow or table not initialized when updating {label}")
        except RuntimeError as e:
            print(f"Error: {e}")
            
    def on_disconnect_clicked(self):
        """Closes the current connection"""
        global currentState
        currentState  = STATE.DISCONNECTED.value
        if self.serial_thread and hasattr(self.serial_thread, 'isRunning') and self.serial_thread.isRunning():
            self.serial_thread.stop()
            self.connection_open = False
            self.config.setEnabled(False)
            self.test.setEnabled(False)
            self.calibrate_ai.setEnabled(False)
            self.exit.setEnabled(False)
            self.statusbar.showMessage("Disconnected")

            try:
                if self.programWindow is not None:
                    self.programWindow.close()
                    self.programWindow = None
                elif self.configWindow is not None:
                    self.configWindow.close()
                    self.configWindow = None
                elif self.calibrateAIWindow is not None:
                    self.calibrateAIWindow.close()
                    self.calibrateAIWindow = None
                elif self.testWindow is not None:
                    self.testWindow.close()  
                    self.testWindow = None  
                elif self.informationwindow is not None:
                    self.informationwindow.close()
                    self.informationwindow = None
                elif self.terminalWindow is not None:
                    self.terminalWindow.close()
                    self.terminalWindow = None
                else:
                    pass
            except RuntimeError as e:
                print(e)

        elif self.serial_thread:
            self.connection_open = False
            self.connection_open = False
            self.config.setEnabled(False)
            self.test.setEnabled(False)
            self.calibrate_ai.setEnabled(False)
            self.exit.setEnabled(False)
            self.statusbar.showMessage("Disconnected")


            try:
                if self.programWindow is not None:
                    self.programWindow.close()
                    self.programWindow = None
                elif self.configWindow is not None:
                    self.configWindow.close()
                    self.configWindow = None
                elif self.calibrateAIWindow is not None:
                    self.calibrateAIWindow.close()
                    self.calibrateAIWindow = None
                elif self.testWindow is not None:
                    self.testWindow.close()  
                    self.testWindow = None  
                elif self.informationwindow is not None:
                    self.informationwindow.close()
                    self.informationwindow = None
                elif self.terminalWindow is not None:
                    self.terminalWindow.close()
                    self.terminalWindow = None
                else:
                    pass
            except RuntimeError as e:
                print(e)
        else:
            self.show_warning_message("Warning", "No active connection to disconnect.")

    def show_warning_message(self, title, message):
        msg_box = HandPointerMessageBox()
        msg_box.setWindowTitle(title)
        msg_box.setWindowIcon(self.window_icon)
        msg_box.setText(message)
        msg_box.setIcon(QMessageBox.Icon.Warning)
        msg_box.exec()

    def show_DataWindow(self):
        self.showdataWindow = ShowDataWindow(self.service)
        self.setCentralWidget(self.showdataWindow)
        if self.terminalWindow is not None:
            self.terminalWindow.close()
            self.terminalWindow = None

    def programFW(self):
        self.programWindow = ProgramWindow(self.window_icon, self.image_load, self.statusbar)
        self.setCentralWidget(self.programWindow)
        if self.terminalWindow is not None:
            self.terminalWindow.deleteLater()
            self.terminalWindow = None

    def configureDevice(self):
        global currentState
        currentState = STATE.CONFIGUREMODE.value
        if self.configWindow is None:
            # Send the appropriate command to the serial thread
            self.serial_thread.send_data('2' + "/n")

            # Create a new instance of ConfigWindow
            self.configWindow = ConfigWindow(self.window_icon, self.serial_thread, self.service)
            
            # Show the ConfigWindow
            self.setCentralWidget(self.configWindow)
            
            # Show status message
            self.statusbar.showMessage("Entered into the Configuration Mode")
        else:
            # If ConfigWindow already exists, simply set it as the central widget
            self.serial_thread.send_data('2' + "/n")
            self.statusbar.showMessage("Entered into the Configuration Mode")
            self.configWindow = ConfigWindow(self.window_icon, self.serial_thread, self.service)
            self.setCentralWidget(self.configWindow)

        if self.terminalWindow is not None:
            self.terminalWindow.deleteLater()
            self.terminalWindow = None

    def testDevice(self):
        global currentState
        currentState = STATE.TESTMODE.value
        if self.testWindow is None:
            # Create a new instance of TestWindow
            self.serial_thread.send_data('1' + "\n")
            self.testWindow = TestWindow(self.terminalWindow, self.serial_thread, self.image_load, self.statusbar,parent=self)
            # Set the TestWindow as the central widget
            self.setCentralWidget(self.testWindow)
            self.statusbar.showMessage("Entered into the Test Mode")
        else:
            self.serial_thread.send_data('1' + "\n")
            self.statusbar.showMessage("Entered into the Test Mode")
            self.testWindow = TestWindow(self.terminalWindow, self.serial_thread, self.image_load, self.statusbar,parent=self)

            # Set the TestWindow as the central widget
            self.setCentralWidget(self.testWindow)


        if self.terminalWindow is not None:
            self.terminalWindow.deleteLater()
            self.terminalWindow = None

    def openTerminalWindow(self):
        if self.terminalWindow is None:
            self.terminalWindow = TerminalWindow()
        self.setCentralWidget(self.terminalWindow)

    def calibrate_AI(self):
        global currentState
        currentState = STATE.CALIBRATEAI.value
        if  self.calibrateAIWindow is None:
            self.serial_thread.send_data('3' + "\n")
            self.calibrateAIWindow = CalibrateAIWindow(self.statusbar, self.window_icon, self.serial_thread)
            self.setCentralWidget(self.calibrateAIWindow)
            self.statusbar.showMessage("Entered into the Calibration AI Mode")
        else:
            self.serial_thread.send_data('3' + "\n")
            self.statusbar.showMessage("Entered into the Calibration AI Mode")
            self.calibrateAIWindow = CalibrateAIWindow(self.statusbar, self.window_icon, self.serial_thread)
            self.setCentralWidget(self.calibrateAIWindow)

        if self.terminalWindow is not None:
            self.terminalWindow.deleteLater()
            self.terminalWindow = None

    def exit_All(self):
        global currentState
        currentState = STATE.EXITNORMALLY.value
        self.serial_thread.send_data('4' + "\n")
        self.config.setEnabled(False)
        self.test.setEnabled(False)
        self.calibrate_ai.setEnabled(False)
        self.exit.setEnabled(False)

        self.statusbar.showMessage(self.data)

        try:
            if self.programWindow is not None:
                self.programWindow.close()
                self.programWindow = None
            elif self.configWindow is not None:
                self.configWindow.close()
                self.configWindow = None
            elif self.calibrateAIWindow is not None:
                self.calibrateAIWindow.close()
                self.calibrateAIWindow = None
            elif self.testWindow is not None:
                self.testWindow.close()  
                self.testWindow = None  
            elif self.informationwindow is not None:
                self.informationwindow.close()
                self.informationwindow = None
            else:
                pass
        except RuntimeError as e:
            print(e)


class InformationWindow(QWidget):
    """Information Window class for displaying information about the DataLogger"""
    def __init__(self):
        super().__init__()
        try:
            self.setContentsMargins(0, 0, 172, 172)
            self.information_layout = QGridLayout()

            self.table = QTableWidget()
            self.table.setColumnCount(2)
            self.table.setHorizontalHeaderLabels(("Item", "Value"))
            self.table.horizontalHeader().setStyleSheet("font-weight: bold")

            self.table.verticalHeader().setVisible(False)
            self.table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
            self.table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
            self.table.viewport().setAutoFillBackground(False)

            self.table.horizontalHeader().setStretchLastSection(True)

            self.information_layout.addWidget(self.table, 0, 0)

            self.predefined_values = ["Serial No", "Model No", "Firmware Version"]
            for i, value in enumerate(self.predefined_values):
                self.table.insertRow(i)
                item = QTableWidgetItem(value)
                item.setFlags(Qt.ItemFlag.ItemIsEnabled)  # Make cells read-only
                self.table.setItem(i, 0, item)

            self.table.resizeRowsToContents()

            self.setLayout(self.information_layout)
        except AttributeError as e:
            print(e)


class ShowDataWindow(QWidget):
    """ShowDataWindow class for displaying data from the Google Cloud Sheet"""

    data_loaded = pyqtSignal(list, list)  # Signal to indicate data has been loaded

    def __init__(self, service):
        super().__init__()
        self.service = service

        self.showdata_layout = QGridLayout()

        self.searchbox = QLineEdit()
        self.searchbox.setPlaceholderText("Search")
        self.searchbox.setStyleSheet("QLineEdit {background-color: white; border: 2px solid gray; border-radius: 10px; padding: 0 8px; }")
        self.searchbox.setFixedSize(250, 30)
        self.showdata_layout.addWidget(self.searchbox, 0, 0)

        self.searchbutton = QPushButton("Search")
        self.searchbutton.setStyleSheet(
            """
            QPushButton {
                background-color: white;
                border: None;
                border-radius: 15px; 
                padding: 8px 16px;
                font-size: 12px;
            }

            QPushButton:hover {
                background-color: #A6F1F4; 
                border-color: grey; 
            }
            """
        )
        self.searchbutton.clicked.connect(self.search_data)

        self.showdata_layout.addWidget(self.searchbutton, 0, 1)

        self.exportbutton = QPushButton("Export Data")
        self.exportbutton.setStyleSheet(
            """
            QPushButton {
                background-color: white;
                border: None;
                border-radius: 15px; 
                padding: 8px 16px;
                font-size: 12px;
            }

            QPushButton:hover {
                background-color: #A6F1F4; 
                border-color: grey; 
            }
            """
        )
        self.exportbutton.clicked.connect(self.export_data)

        self.showdata_layout.addWidget(self.exportbutton, 0, 2)

        self.showdata_table = QTableWidget()
        self.showdata_table.horizontalHeader().setStyleSheet("font-weight: bold")

        self.showdata_layout.addWidget(self.showdata_table, 1, 0, 1, 3)

        self.setLayout(self.showdata_layout)

        self.data_loaded.connect(self.populate_table)

        self.load_data_thread = threading.Thread(target=self.showData_IntoTable)
        self.load_data_thread.start()

    def showData_IntoTable(self):
        """Show data from the Google Cloud Sheet"""
        self.sheet = self.service.spreadsheets()
        self.result = self.sheet.values().get(spreadsheetId=DATABASE_SPREADSHEET_ID, range=DATABASE_RANGE_NAME).execute()
        self.values = self.result.get('values', [])

        self.headers = self.values[0]
        self.data_values_fromsheet = self.values[1:]

        self.data_loaded.emit(self.headers, self.data_values_fromsheet)

    @pyqtSlot(list, list)
    def populate_table(self, headers, data_values_fromsheet):
        for index, value in enumerate(headers):
            self.showdata_table.insertColumn(index)
            self.showdata_table.setHorizontalHeaderItem(index, QTableWidgetItem(value))
            self.showdata_table.resizeColumnToContents(index)

        for i, row_values in enumerate(data_values_fromsheet):
            self.showdata_table.insertRow(i)
            for j, cell_value in enumerate(row_values):
                item = QTableWidgetItem(cell_value)
                item.setFlags(Qt.ItemFlag.ItemIsEnabled)  # Make cells read-only
                self.showdata_table.setItem(i, j, item)

    def search_data(self):
        """Search data into the Table"""
        search_text = self.searchbox.text()
        # print("Search Text:", search_text)  # Debugging statement

        items = self.showdata_table.findItems(search_text, Qt.MatchFlag.MatchFixedString)
        # print("Number of Items Found:", len(items))  # Debugging statement

        if not items:
            print("No items found.")
            return

        # Clear previous selections and reset background color
        for row in range(self.showdata_table.rowCount()):
            for column in range(self.showdata_table.columnCount()):
                item = self.showdata_table.item(row, column)
                if item:
                    item.setSelected(False)
                    item.setBackground(QBrush())

        for item in items:
            # print("Row:", item.row(), "Column:", item.column())  # Debugging statement
            for column in range(self.showdata_table.columnCount()):
                selected_item = self.showdata_table.item(item.row(), column)
                selected_item.setSelected(True)
                selected_item.setBackground(QBrush(QColor("#2C7FEC")))  

    def export_data(self):
        """Export data from the table and make Excel file from it"""
        filename = "ConfigurationData.xlsx"
        headers = []
        for col in range(self.showdata_table.columnCount()):
            header_item = self.showdata_table.horizontalHeaderItem(col)
            if header_item is not None:
                headers.append(header_item.text())
        all_rows = []
        for row in range(self.showdata_table.rowCount()):
            row_values = []
            for column in range(self.showdata_table.columnCount()):
                item = self.showdata_table.item(row, column)
                if item is not None:
                    row_values.append(item.text())
                else:
                    row_values.append('')
            all_rows.append(row_values)
        
        all_data = [headers] + all_rows
        # print(all_data)  # Replace this with the code to export data
        self.write_into_excel(filename, all_data)

    def write_into_excel(self, filename, all_data):
        try:
            workbook = load_workbook(filename)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active

        # Extract existing data
        existing_data = [list(row) for row in sheet.iter_rows(values_only=True)]
        
        # Write data to the Excel sheet if it doesn't already exist
        for row_data in all_data:
            if row_data not in existing_data:
                sheet.append(row_data)
        
        # Save the workbook
        workbook.save(filename)


class AboutDialog(QMessageBox):
    def  __init__(self, image_load):
        super().__init__()
        self.image_load = image_load

        self.setWindowTitle("About")
        content = """
HTCU: HRMS Testing & Config Utility Software
Used for Uploading Firmware, Testing & Config the device
Version: HTCU-V1.0.1
Developed By: Holmium technologies Pvt Ltd
Released Date: 03/04/2024

Copyright © 2023 Holmium Technologies. All Rights Reserved
"""
        self.setText(content)
        self.setIcon(QMessageBox.Icon.Information)
        window_icon = QIcon(self.image_load.load_image("icon\logo.png").scaled(60, 60))
        self.setWindowIcon(window_icon)


class ProgramWindow(QWidget):
    def __init__(self, window_icon, image_load, statusbar):
        super().__init__()
        self.image_load = image_load
        self.statusbar = statusbar 
        self.window_icon = window_icon
        self.successs_message ="Success"
        self.error_message = "Error"

        self.selected_file_paths = []

        # Load selected file paths, checkboxes, and deleted checkboxes from settings
        settings = QSettings("MyCompany", "MyApp")
        self.selected_file_paths = settings.value("selected_file_paths", [])
        self.checkbox_states = settings.value("checkbox_states", {})  # Dictionary to store checkbox states
        self.deleted_checkboxes = settings.value("deleted_checkboxes", [])

        self.grid_layout = QGridLayout()

        select_file = QLabel("Select File to Upload:")
        self.filename_edit = QLineEdit()
        self.filename_edit.setFixedSize(270, 30)
        self.filename_edit.setStyleSheet("QLineEdit {background-color: white; border: 2px solid gray; border-radius: 10px; padding: 0 8px; }")
        self.browse_button = QPushButton("Browse_File")
        self.browse_button.setStyleSheet(
            """
            QPushButton {
                background-color: white;
                border: None;
                border-radius: 15px; 
                padding: 8px 16px;
                font-size: 12px;
            }

            QPushButton:hover {
                background-color: #A6F1F4; 
                border-color: grey; 
            }
            """
        )
        self.browse_button.setFixedSize(95, 30)

        self.browse_button.clicked.connect(self.select_file)

        self.grid_layout.addWidget(select_file, 0, 0)
        self.grid_layout.addWidget(self.filename_edit, 0, 1)
        self.grid_layout.addWidget(self.browse_button, 0, 2)

        self.upload = QPushButton("Upload")
        self.upload.setStyleSheet(
            """
            QPushButton {
                background-color: white;
                border: None;
                border-radius: 15px; 
                padding: 8px 16px;
                font-size: 12px;
            }

            QPushButton:hover {
                background-color: #A6F1F4; 
                border-color: grey; 
            }
            """
        )
        self.upload.setFixedSize(90, 30)

        self.grid_layout.addWidget(self.upload, 0, 3)
        self.upload.clicked.connect(self.upload_program)

        self.text_area = QTextEdit(readOnly=True)
        self.text_area.setStyleSheet("QTextEdit {background-color: white; border: 2px solid gray; border-radius: 10px; padding: 0 8px; }")
        self.grid_layout.addWidget(self.text_area, 1, 0,  1, 4)

        self.upload_thread = None
        self.setLayout(self.grid_layout)

        # Recreate checkboxes when GUI is initialized
        self.recreate_checkboxes()

    def recreate_checkboxes(self):
        # Clear existing checkboxes
        for i in reversed(range(self.grid_layout.count())):
            widget = self.grid_layout.itemAt(i).widget()
            if isinstance(widget, QCheckBox):
                self.grid_layout.removeWidget(widget)
                widget.deleteLater()

        for index, file_path in enumerate(self.selected_file_paths):
            if file_path not in self.deleted_checkboxes:
                checkbox = DeleteableCheckBox(file_path)
                # Connect the deleteRequested signal to the delete_checkbox method
                checkbox.deleteRequested.connect(lambda checkbox=checkbox: self.delete_checkbox(checkbox))
                # Set checkbox state based on the saved states
                checkbox.setChecked(self.checkbox_states.get(file_path, False))

                checkbox.stateChanged.connect(lambda state, checkbox=checkbox: self.update_selected_file(checkbox))

                self.grid_layout.addWidget(checkbox, index + 3, 0, 1, 3)
                # self.checkboxes.append(checkbox)  # Add checkbox to the list for later reference


    def select_file(self):
        file_dialog = QFileDialog(self)
        file_dialog.setWindowTitle("Select a File")
        file_dialog.setFileMode(QFileDialog.FileMode.ExistingFile)

        if file_dialog.exec():
            selected_file = file_dialog.selectedFiles()
            file_path = selected_file[0].replace("/", "\\")

            self.filename_edit.setText(file_path)

            # Add the selected file path to the list
            self.selected_file_paths.append(file_path)

            # Save selected file paths to settings
            self.save_settings()

            # Recreate checkboxes with updated file paths
            self.recreate_checkboxes()

    def upload_program(self):
        self.statusbar.clearMessage()
        port = self.parent().comboBox.currentText()
        if not port:
            msg_box = HandPointerMessageBox()
            msg_box.setWindowTitle("Warning")
            msg_box.setWindowIcon(self.window_icon)
            msg_box.setText("Please select a port.")
            msg_box.setIcon(QMessageBox.Icon.Warning)
            msg_box.exec()
            return

        filename = str(self.filename_edit.text()).replace('\\', '\\\\')
        if not filename:
            msg_box = HandPointerMessageBox()
            msg_box.setWindowTitle("Warning")
            msg_box.setWindowIcon(self.window_icon)
            msg_box.setText("Please select a file.")
            msg_box.setIcon(QMessageBox.Icon.Warning)
            msg_box.exec()
            return

        else:
            # Command to execute
            command = f'esptool --port {port} write_flash 0x0000 {filename}'

            # Start the upload in a separate thread
            self.upload_thread = UploadThread(command)
            self.upload_thread.output_received.connect(self.update_text_area)
            self.upload_thread.start()

    def update_text_area(self, line):
        error_list = ["A fatal error occurred: Failed to connect to Espressif device: No serial data received.",
                      "A serial exception error occurred: Cannot configure port, something went wrong. Original message: PermissionError(13, 'Access is denied.', None, 5)", \
                      "A fatal error occurred: The chip stopped responding.", "A serial exception error occurred: Write timeout", \
                      "A fatal error occurred: Failed to connect to Espressif device: Download mode successfully detected, but getting no sync reply: The serial TX path seems to be down.", \
                      "A fatal error occurred: Failed to connect to Espressif device: Serial data stream stopped: Possible serial noise or corruption."]

        # Append the output line to the QTextEdit box
        self.text_area.append(line)
        # Create QLabel for icon and message if they don't exist
        if not hasattr(self, 'icon_label'):
            self.icon_label = QLabel()
            self.message_label = QLabel()
            self.statusbar.addWidget(self.icon_label)
            self.statusbar.addWidget(self.message_label)

        # Update icon and message based on the content of text_data
        if "Connecting...." in line:
            self.image = self.image_load.load_image("icon\icons8-waiting-50.png").scaled(30, 30)

            self.icon_label.setPixmap(self.image)
            self.message_label.setText("Writing...")
            self.message_label.setStyleSheet("color: darkblack; font-family: times; font-size: 20px;")

            # Create opacity effect
            self.opacity_effect = QGraphicsOpacityEffect()
            self.opacity_effect.setOpacity(1.0)  # Initial opacity
            self.icon_label.setGraphicsEffect(self.opacity_effect)

            # Create animation
            self.fade_animation = QPropertyAnimation(self.opacity_effect, b"opacity")
            self.fade_animation.setStartValue(1.0)
            self.fade_animation.setEndValue(0.0)
            self.fade_animation.setDuration(1000)  # 1 second duration
            self.fade_animation.setLoopCount(-1)  # Infinite loop
            self.fade_animation.start()

        if "Hash of data verified." in line:
            self.statusbar.removeWidget(self.icon_label)
            self.statusbar.removeWidget(self.message_label)

        elif "Leaving..." in line:

            # Stop the rotation animation
            self.fade_animation.stop()

            self.success_image = self.image_load.load_image("icon\icons8-success-94.png").scaled(30, 30)
            self.icon = self.show_temporary_image(self.success_image, self.successs_message, duration=3000)

        for error in error_list:
            if error in line:
    
                self.error_image = self.image_load.load_image("icon\icons8-error-94.png").scaled(30, 30)
                self.icon = self.show_temporary_image(self.error_image, self.error_message, duration=3000)

    def show_temporary_image(self, image, message, duration):
        """Show an image and then hide it after a specified duration."""

        icon_label = QLabel()
        message_label = QLabel()
        icon_label.setPixmap(image)
        message_label.setText(message)
        if message == self.successs_message:
            message_label.setStyleSheet("color: green; font-family: times; font-size: 20px;")
        elif message == self.error_image:
            message_label.setStyleSheet("color: red; font-family: times; font-size: 20px;")
        else:
            pass

        self.statusbar.addWidget(icon_label)
        self.statusbar.addWidget(message_label)

        # Use a QTimer to remove the label after the specified duration
        timer = QTimer(self)
        timer.singleShot(duration, lambda: self.statusbar.removeWidget(icon_label))
        timer.singleShot(duration, lambda: self.statusbar.removeWidget(message_label))
        timer.start()

        return icon_label


    def save_settings(self):
        # Save selected file paths, checkbox states, and deleted checkboxes to settings
        settings = QSettings("MyCompany", "MyApp")
        settings.setValue("selected_file_paths", self.selected_file_paths)
        # settings.setValue("deleted_checkboxes", self.deleted_checkboxes)

    def delete_checkbox(self, checkbox):
       # Remove the checkbox from the layout
        self.grid_layout.removeWidget(checkbox)
        checkbox.deleteLater()

        # Remove the checkbox from selected_file_paths if it exists
        file_path = checkbox.text()
        if file_path in self.selected_file_paths:
            self.selected_file_paths.remove(file_path)

        # Save settings
        self.save_settings()

    def update_selected_file(self, checkbox):
        if checkbox.isChecked():
            self.filename_edit.setText(checkbox.text())
        else:
            if self.filename_edit.text() == checkbox.text():
                self.filename_edit.clear()

class DeleteableCheckBox(QCheckBox):
    deleteRequested = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.delete_image = ImageLoader()
        self.deleteIcon = self.delete_image.load_image("icon\icons8-delete-16.png")

    def paintEvent(self, event):
        # Call the base class paintEvent to draw the checkbox
        super().paintEvent(event)
        painter = QPainter(self)

        if self.isChecked():
            # Draw the delete icon if the checkbox is checked
            deleteIconSize = 16
            deleteIconX = self.width() - deleteIconSize - 4
            deleteIconY = (self.height() - deleteIconSize) / 2
            deleteIconRect = QRect(int(deleteIconX), int(deleteIconY), deleteIconSize, deleteIconSize)
            painter.drawPixmap(deleteIconRect, self.deleteIcon)

    def mousePressEvent(self, event):
        if self.isChecked():
            # Check if the mouse click is within the delete icon area
            deleteIconSize = 16
            deleteIconX = self.width() - deleteIconSize - 4
            deleteIconY = (self.height() - deleteIconSize) / 2

            if event.button() == Qt.MouseButton.LeftButton and \
                    deleteIconX <= event.pos().x() <= deleteIconX + deleteIconSize and \
                    deleteIconY <= event.pos().y() <= deleteIconY + deleteIconSize:
                # Emit signal for delete action if the delete icon is clicked
                self.deleteRequested.emit()
                return

        # Call the base class mousePressEvent for default behavior
        super().mousePressEvent(event)
            
            
class ConfigWindow(QWidget):
    def  __init__(self, window_icon, serial_thread, service):
        super().__init__()
        self.serial_thread = serial_thread
        self.window_icon = window_icon
        self.service = service

        # self.setContentsMargins(350, 350, 350, 350)
        
        layout = QGridLayout()

        device_name = QLabel("HRMS-E32:")
        self.device_combo = QComboBox()
        self.device_combo.setCursor(Qt.CursorShape.PointingHandCursor)
        self.device_combo.setStyleSheet("QComboBox {background-color: white; border: 2px solid gray; border-radius: 10px; padding: 0 8px; }")
        self.device_combo.setPlaceholderText("Select Device")
        devicename = ["W", "WGL", "WG", "WL", "G", "L", "GL", "PG", "PP", "PE"]
        self.device_combo.addItems(devicename)

        layout.addWidget(device_name, 0, 0)
        layout.addWidget(self.device_combo, 0, 1)

        # self.subdevice_combo = QComboBox()
        # self.subdevice_combo.setStyleSheet("background-color: white;")
        # self.subdevice_combo.addItems(["Zero Export", "PVDG", "Both"])
        # layout.addWidget(self.subdevice_combo, 0, 2)

        # self.device_combo.currentIndexChanged.connect(self.update_subdevices)
        # self.update_subdevices(self.device_combo.currentIndex())

        serial_number = QLabel("Serial No.")
        self.serial_no = QComboBox()
        self.serial_no.setStyleSheet("QComboBox {background-color: white; border: 2px solid gray; border-radius: 10px; padding: 0 8px; }")
        self.serial_no.setPlaceholderText("Enter Serial Number")
        layout.addWidget(serial_number, 1, 0)
        layout.addWidget(self.serial_no, 1, 1)

        threading.Thread(target=self.add_serialno()).start()

        pcb_partnumber = QLabel("PCB Part Number")
        self.pcb_number = QComboBox()
        self.pcb_number.setStyleSheet("QComboBox {background-color: white; border: 2px solid gray; border-radius: 10px; padding: 0 8px; }")
        self.pcb_number.setPlaceholderText("Enter PCB Part Number")
        layout.addWidget(pcb_partnumber, 2, 0)
        layout.addWidget(self.pcb_number, 2, 1)

        threading.Thread(target=self.add_pcbpartnumber()).start()

        gsm_partnumber = QLabel("GSM Part Number")
        self.gsm_number = QComboBox()
        self.gsm_number.setStyleSheet("QComboBox {background-color: white; border: 2px solid gray; border-radius: 10px; padding: 0 8px; }")
        self.gsm_number.setPlaceholderText("Enter GSM Part Number")
        layout.addWidget(gsm_partnumber, 3, 0)
        layout.addWidget(self.gsm_number, 3, 1)

        threading.Thread(target=self.add_gsmpartnumber()).start()

        ethernet_partnumber = QLabel("Ethernet Part Number")
        self.ethernet_number = QComboBox()
        self.ethernet_number.setStyleSheet("QComboBox {background-color: white; border: 2px solid gray; border-radius: 10px; padding: 0 8px; }")
        self.ethernet_number.setPlaceholderText("Enter Ethernet Part Number")
        layout.addWidget(ethernet_partnumber, 4, 0)
        layout.addWidget(self.ethernet_number, 4, 1)

        threading.Thread(target=self.add_ethernetpartnumber()).start()

        password = QLabel("Password")
        self.password = PasswordLineEdit()
        self.password.setStyleSheet("background-color: white")
        # self.password.setPlaceholderText("Enter Password")
        layout.addWidget(password, 5, 0)
        layout.addWidget(self.password, 5, 1)

        configured = QLabel("Configured and Tested By: ")
        self.configured_by = QComboBox()
        self.configured_by.setPlaceholderText("Select Your Name")
        self.configured_by.setStyleSheet("QComboBox {background-color: white; border: 2px solid gray; border-radius: 10px; padding: 0 8px; }")
        Users = ["Satish", "Arun", "Rizwan", "khushbu", "Surya", "Jacob", "Sidharth", "Kishan", "Sandeep"]
        self.configured_by.addItems(Users)
        layout.addWidget(configured, 6, 0)
        layout.addWidget(self.configured_by, 6, 1)

        self.configure = QPushButton("Configure")
        self.configure.setStyleSheet(
                    """
            QPushButton {
                background-color: white;
                border: None;
                border-radius: 15px; 
                padding: 8px 16px;
                font-size: 12px;
            }

            QPushButton:hover {
                background-color: #A6F1F4; 
                border-color: grey; 
            }
            """
        )
        self.configure.setFixedSize(88,30)
        self.configure.setCursor(Qt.CursorShape.PointingHandCursor)
        self.configure.clicked.connect(self.on_configure_clicked)
        layout.addWidget(self.configure,  7, 3)

        self.setLayout(layout)

    # def update_subdevices(self, index):
    #     if index == 0:
    #         self.subdevice_combo.hide()
    #     elif index == 1:
    #         self.subdevice_combo.show()
    #     elif index == 2:
    #         self.subdevice_combo.hide()
    #     elif index == 3:
    #         self.subdevice_combo.hide()

    def add_serialno(self):
        try:
            self.sheet = self.service.spreadsheets()
            self.result = self.sheet.values().get(spreadsheetId=DATABASE_SPREADSHEET_ID, range=DATABASE_RANGE_NAME1).execute()
            self.values_serial = self.result.get('values', [])
            
            if not self.values_serial:
                print("No data found in the specified range.")
                return
            
            self.headers = self.values_serial[0]  # Assume the first row is the header
            self.data_rows_serial = self.values_serial[1:]  # The rest are data rows
            # print(self.data_rows_serial)

            for row in self.data_rows_serial:
                try:
                    if int(row[1]) == 0:  # Ensure row[1] is an integer and equals 0
                        self.serial_no.addItem(row[0])
                except ValueError:
                    # Skip rows where row[1] is not an integer
                    continue

        except Exception as e:
            print(f"Error fetching data from Google Sheets: {e}")

    def add_pcbpartnumber(self):
        try:
            # sheet = self.service.spreadsheets()
            self.result = self.sheet.values().get(spreadsheetId=DATABASE_SPREADSHEET_ID, range=DATABASE_RANGE_NAME2).execute()
            values = self.result.get('values', [])
            
            if not values:
                print("No data found in the specified range.")
                return
            
            headers = values[0]  # Assume the first row is the header
            # print(headers)
            self.data_rows_pcb = values[1:]  # The rest are data rows

            for row in self.data_rows_pcb:
                try:
                    if int(row[1]) == 0:  # Ensure row[1] is an integer and equals 0
                        self.pcb_number.addItem(row[0])
                except ValueError:
                    # Skip rows where row[1] is not an integer
                    continue

        except Exception as e:
            print(f"Error fetching data from Google Sheets: {e}")

    def add_gsmpartnumber(self):
        try:
            # sheet = self.service.spreadsheets()
            result = self.sheet.values().get(spreadsheetId=DATABASE_SPREADSHEET_ID, range=DATABASE_RANGE_NAME3).execute()
            values = result.get('values', [])
            
            if not values:
                print("No data found in the specified range.")
                return
            
            headers = values[0]  # Assume the first row is the header
            self.data_rows_gsm = values[1:]  # The rest are data rows

            for row in self.data_rows_gsm:
                try:
                    if int(row[1]) == 0:  # Ensure row[1] is an integer and equals 0
                        self.gsm_number.addItem(row[0])
                except ValueError:
                    # Skip rows where row[1] is not an integer
                    continue

        except Exception as e:
            print(f"Error fetching data from Google Sheets: {e}")

    def add_ethernetpartnumber(self):
        try:
            # sheet = self.service.spreadsheets()
            result = self.sheet.values().get(spreadsheetId=DATABASE_SPREADSHEET_ID, range=DATABASE_RANGE_NAME4).execute()
            values = result.get('values', [])
            
            if not values:
                print("No data found in the specified range.")
                return
            
            headers = values[0]  # Assume the first row is the header
            self.data_rows_ethernet = values[1:]  # The rest are data rows

            for row in self.data_rows_ethernet:
                try:
                    if int(row[1]) == 0:  # Ensure row[1] is an integer and equals 0
                        self.ethernet_number.addItem(row[0])
                except ValueError:
                    # Skip rows where row[1] is not an integer
                    continue

        except Exception as e:
            print(f"Error fetching data from Google Sheets: {e}")
        
    def on_configure_clicked(self):
        # print("configure clicked")
        global currentState
        currentState = STATE.CONFIGBUTTON.value
        if not self.device_combo.currentText():
            msg_box = HandPointerMessageBox()
            msg_box.setWindowTitle("Warning")
            msg_box.setWindowIcon(self.window_icon)
            msg_box.setText("Incorrect Model No. Please Select the Model No."   )
            msg_box.setIcon(QMessageBox.Icon.Warning)
            msg_box.exec()
            return

        # serial_pattern = re.compile(r'^((HO-[\w\d]+$)|(HRMS-E32-\d+$))')
        # if not serial_pattern.match(self.serial_no.text()):
        #     msg_box = HandPointerMessageBox()
        #     msg_box.setWindowTitle("Warning")
        #     msg_box.setWindowIcon(self.window_icon)
        #     msg_box.setText("Serial No. format should be HO/HRMS- followed by Character or Digits.")
        #     msg_box.setIcon(QMessageBox.Icon.Warning)
        #     msg_box.exec()
        #     return

        if not self.serial_no.currentText():
            msg_box = HandPointerMessageBox()
            msg_box.setWindowTitle("Warning")
            msg_box.setWindowIcon(self.window_icon)
            msg_box.setText("Incorrect Serial No. Please Select the Serial No."   )
            msg_box.setIcon(QMessageBox.Icon.Warning)
            msg_box.exec()
            return
        
        if not self.pcb_number.currentText():
            msg_box = HandPointerMessageBox()
            msg_box.setWindowTitle("Warning")
            msg_box.setWindowIcon(self.window_icon)
            msg_box.setText("Incorrect PCB part number. Please Select the PCB part No."   )
            msg_box.setIcon(QMessageBox.Icon.Warning)
            msg_box.exec()
            return
        
        if not self.gsm_number.currentText():
            msg_box = HandPointerMessageBox()
            msg_box.setWindowTitle("Warning")
            msg_box.setWindowIcon(self.window_icon)
            msg_box.setText("Incorrect GSM part number. Please Select the GSM part No."   )
            msg_box.setIcon(QMessageBox.Icon.Warning)
            msg_box.exec()
            return
        
        if not self.ethernet_number.currentText():
            msg_box = HandPointerMessageBox()
            msg_box.setWindowTitle("Warning")
            msg_box.setWindowIcon(self.window_icon)
            msg_box.setText("Incorrect Ehternet part number. Please Select the Ethernet part No."   )
            msg_box.setIcon(QMessageBox.Icon.Warning)
            msg_box.exec()
            return


        # Validate Password
        if self.password.password_edit.text() != "HO-1810":
            msg_box = HandPointerMessageBox()
            msg_box.setWindowTitle("Warning")
            msg_box.setWindowIcon(self.window_icon)
            msg_box.setText("Incorrect password. Please enter the correct password." )
            msg_box.setIcon(QMessageBox.Icon.Warning)
            msg_box.exec()
            return
        
        if not self.configured_by.currentText():
            msg_box = HandPointerMessageBox()
            msg_box.setWindowTitle("Warning")
            msg_box.setWindowIcon(self.window_icon)
            msg_box.setText("Please Select Your Name.")
            msg_box.setIcon(QMessageBox.Icon.Warning)
            msg_box.exec()
            return
        
        """Save data from GUI elements into Google SPreadSheet"""
        # filename = "ConfigurationData.xlsx"
        timestamp = datetime.datetime.now()
        # Format timestamp as string
        formatted_timestamp = timestamp.strftime("%Y-%m-%d %H:%M:%S")
        self.serial_number = self.serial_no.currentText()
        device_type = self.device_combo.currentText()
        testing = "OK"
        configured = self.configured_by.currentText()
        self.pcb_number = self.pcb_number.currentText()
        self.gsm_number = self.gsm_number.currentText()
        self.ethernet_number = self.ethernet_number.currentText()

        data = [[formatted_timestamp, self.serial_number, device_type, testing, configured, self.pcb_number, self.gsm_number, self.ethernet_number]]

        threading.Thread(target = self.write_into_GoogleSheet, args = (data,)).start()
        
        # Perform time-consuming operations in a separate thread
        threading.Thread(target=self.send_configuration_data).start()
        # time.sleep(10)

    def send_configuration_data(self):
        # if "Enter New Serial No" in self.parent().data:
        serial_number = self.serial_no.currentText() + "\n"
        self.serial_thread.send_data(serial_number)
        time.sleep(2)
            # pass

        # elif "Enter Password to Save New Serial No" in self.parent().data:
        password = self.password.password_edit.text() + "\n"
        self.serial_thread.send_data(password)
        time.sleep(4)
            # pass

        # elif "Select model number:" in self.parent().data:
        model_number = str(self.device_combo.currentIndex() + 1) + "\n"
        self.serial_thread.send_data(model_number)

        threading.Thread(target = self.send_data_after_Configuration).start()

    def send_data_after_Configuration(self):
        # if "E" in self.parent().data[:1]:
        time.sleep(5)
        self.serial_thread.send_data("HRMS-1810" + "\n")
            
        global currentState
        currentState = STATE.CONNECTED.value

    def  write_into_GoogleSheet(self, data):
        """Saving data into the Google Sheet"""
        try:
            # self.sheet = self.service.spreadsheets()
            values = data
            body = {'values': values}
            result = self.sheet.values().append(
                spreadsheetId=DATABASE_SPREADSHEET_ID,
                range=DATABASE_RANGE_NAME,
                valueInputOption="USER_ENTERED",
                body=body
            ).execute()
            # print(f"Data successfully written to Google Sheets: {result}")
        except Exception as e:
            print(f"Error writing to Google Sheets: {e}")


        """Updating the Status flag of Serial Number Google sheets"""
        # Find the matching row for the serial number and update the second column
        row_to_update = None
        for index, row in enumerate(self.data_rows_serial):
            if row[0] == self.serial_number:
                row_to_update = index + 2  # Adding 2 to account for header row and 0-indexing
                break
        if row_to_update is not None:
            try:
                range_to_update = f"Sheet2!B{row_to_update}"  # Adjust the sheet name and column as necessary
                value_input_option = "RAW"
                value_range_body = {
                    "values": [[1]]  # The value to update the cell with
                }
                self.sheet.values().update(
                    spreadsheetId=DATABASE_SPREADSHEET_ID,
                    range=range_to_update,
                    valueInputOption=value_input_option,
                    body=value_range_body
                ).execute()
                # print(f"Data successfully updated to Serial Number Google Sheets: {result}")
            except Exception as e:
                print(f"Error updating to google sheets: {e}")
                

        """Updating the Status flag of PCB Part Number Google sheets"""
        pcbrow_to_update = None
        for index, row in enumerate(self.data_rows_pcb):
            if row[0] == self.pcb_number:
                pcbrow_to_update = index + 2  # Adding 2 to account for header row and 0-indexing
                break
        if pcbrow_to_update is not None:
            try:
                pcbrange_to_update = f"Sheet3!B{pcbrow_to_update}"  # Adjust the sheet name and column as necessary
                pcbvalue_input_option = "RAW"
                pcbvalue_range_body = {
                    "values": [[1]]  # The value to update the cell with
                }
                self.sheet.values().update(
                    spreadsheetId=DATABASE_SPREADSHEET_ID,
                    range=pcbrange_to_update,
                    valueInputOption=pcbvalue_input_option,
                    body=pcbvalue_range_body
                ).execute()
                # print(f"Data successfully updated to PCB Part Number Google Sheets: {result}")
            except Exception as e:
                print(f"Error updating to google sheets: {e}")

        
        """Updating the Status flag of GSM Part Number Google sheets"""
        gsmrow_to_update = None
        for index, row in enumerate(self.data_rows_gsm):
            if row[0] == self.gsm_number:
                gsmrow_to_update = index + 2  # Adding 2 to account for header row and 0-indexing
                break
        if gsmrow_to_update is not None:
            try:
                gsmrange_to_update = f"Sheet4!B{gsmrow_to_update}"  # Adjust the sheet name and column as necessary
                gsmvalue_input_option = "RAW"
                gsmvalue_range_body = {
                    "values": [[1]]  # The value to update the cell with
                }
                self.sheet.values().update(
                    spreadsheetId=DATABASE_SPREADSHEET_ID,
                    range=gsmrange_to_update,
                    valueInputOption=gsmvalue_input_option,
                    body=gsmvalue_range_body
                ).execute()
                # print(f"Data successfully updated to GSM Part Number Google Sheets: {result}")
            except Exception as e:
                print(f"Error updating to google sheets: {e}")

        
        """Updating the Status flag of Ethernet Part Number Google sheets"""
        ethernetrow_to_update = None
        for index, row in enumerate(self.data_rows_ethernet):
            if row[0] == self.ethernet_number:
                ethernetrow_to_update = index + 2  # Adding 2 to account for header row and 0-indexing
                break
        if ethernetrow_to_update is not None:
            try:
                ethernetrange_to_update = f"Sheet5!B{ethernetrow_to_update}"  # Adjust the sheet name and column as necessary
                ethernetvalue_input_option = "RAW"
                ethernetvalue_range_body = {
                    "values": [[1]]  # The value to update the cell with
                }
                self.sheet.values().update(
                    spreadsheetId=DATABASE_SPREADSHEET_ID,
                    range=ethernetrange_to_update,
                    valueInputOption=ethernetvalue_input_option,
                    body=ethernetvalue_range_body
                ).execute()
                # print(f"Data successfully updated to Ethernet Part Number Google Sheets: {result}")
            except Exception as e:
                print(f"Error updating to google sheets: {e}")
            

class PasswordLineEdit(QWidget):
    def __init__(self):
        super().__init__()

        self.password_image = ImageLoader()
        
        self.password_edit = QLineEdit()
        self.password_edit.setStyleSheet("QLineEdit {background-color: white; border: 2px solid gray; border-radius: 10px; padding: 0 8px; }")
        self.password_edit.setText("HO-1810")
        self.password_edit.setEchoMode(QLineEdit.EchoMode.Password)
        self.password_edit.textChanged.connect(self.toggle_eye_visibility)

        self.show_password_button = QPushButton()
        icon = QIcon(self.password_image.load_image("icon\icons8-eye-24.png"))
        self.show_password_button.setIcon(icon)
        self.show_password_button.setStyleSheet("border: None; padding: 0px; background-color: transparent")
        self.show_password_button.setCheckable(True)
        self.show_password_button.toggled.connect(self.toggle_password_visibility)

        self.timer = QTimer()
        self.timer.setInterval(200)  # Delay in milliseconds
        self.timer.setSingleShot(True)
        self.timer.timeout.connect(self.update_password_echo_mode)

        layout = QHBoxLayout()
        layout.addWidget(self.password_edit)
        layout.addWidget(self.show_password_button)

        self.setLayout(layout)

        # Initially hide the eye icon
        self.show_password_button.setVisible(False)

    def toggle_password_visibility(self):
        self.timer.start()

    def update_password_echo_mode(self):
        if self.show_password_button.isChecked():
            self.password_edit.setEchoMode(QLineEdit.EchoMode.Normal)
        else:
            self.password_edit.setEchoMode(QLineEdit.EchoMode.Password)
    
    def toggle_eye_visibility(self, text):
        """Toggle visibility of the password character"""
        self.show_password_button.setVisible(bool(text)) 


class TestWindow(QWidget):
    """To test the functionalities of the firmware which is uploaded into the hardware."""
    def  __init__(self , terminalWindow, serial_thread, image_load, statusbar, parent = None):
        super().__init__(parent)

        self.statusbar = statusbar
        self.serial_thread = serial_thread
        self.image_load = image_load
        self.terminalWindow = terminalWindow

        # self.setContentsMargins(300, 00, 300, 300)

        layout = QGridLayout()

        self.movie_label = QLabel()
        self.movie = self.image_load.load_gif("icon\icons8-spinner.gif")
        self.movie.setScaledSize(QSize(50, 50))  
        self.movie.setSpeed(100)
        self.movie.start()
        self.movie_label.setMovie(self.movie)
        # self.movie_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.movie_label.setVisible(False)
        self.statusbar.addPermanentWidget(self.movie_label, 0)

    
        # Initialize QLabel for movie_test
        self.movie_testAll = QLabel()
        self.movie_test = self.image_load.load_gif("icon\icons8-spinner.gif")
        self.movie_test.setScaledSize(QSize(50, 50))  # Set a smaller size
        self.movie_test.setSpeed(100)  # Set the frame rate (lower value means slower)
        self.movie_test.start()
        self.movie_testAll.setMovie(self.movie_test)
        # self.movie_testAll.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.movie_testAll.setVisible(False)
        self.statusbar.addPermanentWidget(self.movie_testAll, 0)

        self.testrtc = QPushButton("Test RTC")
        self.testrtc.setStyleSheet(
                    """
            QPushButton {
                background-color: white;
                border: None;
                border-radius: 15px; 
                padding: 8px 16px;
                font-size: 12px;
            }

            QPushButton:hover {
                background-color: #A6F1F4; 
                border-color: grey; 
            }
            """
        )
        self.testrtc.setCursor(Qt.CursorShape.PointingHandCursor)
        self.testrtc.setFixedSize(120, 30)
        layout.addWidget(self.testrtc, 0, 0)
        self.testrtc.clicked.connect(self.test_RTC)

        self.testgsm = QPushButton("Test GSM")
        self.testgsm.setStyleSheet(
                    """
            QPushButton {
                background-color: white;
                border: None;
                border-radius: 15px; 
                padding: 8px 16px;
                font-size: 12px;
            }

            QPushButton:hover {
                background-color: #A6F1F4; 
                border-color: grey; 
            }
            """
        )
        self.testgsm.setCursor(Qt.CursorShape.PointingHandCursor)
        self.testgsm.setFixedSize(120, 30)
        layout.addWidget(self.testgsm, 0, 1)
        self.testgsm.clicked.connect(self.test_GSM)

        self.testwifi = QPushButton("Test WiFi")
        self.testwifi.setStyleSheet(
                    """
            QPushButton {
                background-color: white;
                border: None;
                border-radius: 15px; 
                padding: 8px 16px;
                font-size: 12px;
            }

            QPushButton:hover {
                background-color: #A6F1F4; 
                border-color: grey; 
            }
            """
        )
        self.testwifi.setCursor(Qt.CursorShape.PointingHandCursor)
        self.testwifi.setFixedSize(120, 30)
        layout.addWidget(self.testwifi, 1, 0)
        self.testwifi.clicked.connect(self.test_WiFi)

        self.testethernet = QPushButton("Test Ethernet")
        self.testethernet.setStyleSheet(
                    """
            QPushButton {
                background-color: white;
                border: None;
                border-radius: 15px; 
                padding: 8px 16px;
                font-size: 12px;
            }

            QPushButton:hover {
                background-color: #A6F1F4; 
                border-color: grey; 
            }
            """
        )
        self.testethernet.setCursor(Qt.CursorShape.PointingHandCursor)
        self.testethernet.setFixedSize(120, 30)
        layout.addWidget(self.testethernet, 1, 1)
        self.testethernet.clicked.connect(self.test_Ethernet)

        self.testsdcard = QPushButton("Test SD Card")
        self.testsdcard.setStyleSheet(
                    """
            QPushButton {
                background-color: white;
                border: None;
                border-radius: 15px; 
                padding: 8px 16px;
                font-size: 12px;
            }

            QPushButton:hover {
                background-color: #A6F1F4; 
                border-color: grey; 
            }
            """
        )
        self.testsdcard.setCursor(Qt.CursorShape.PointingHandCursor)
        self.testsdcard.setFixedSize(120, 30)
        layout.addWidget(self.testsdcard, 2, 0)
        self.testsdcard.clicked.connect(self.test_SDCard)

        self.testmodbusrtu = QPushButton("Test Modbus RTU")
        self.testmodbusrtu.setStyleSheet(
                    """
            QPushButton {
                background-color: white;
                border: None;
                border-radius: 15px; 
                padding: 8px 16px;
                font-size: 11px;
            }

            QPushButton:hover {
                background-color: #A6F1F4; 
                border-color: grey; 
            }
            """
        )
        self.testmodbusrtu.setCursor(Qt.CursorShape.PointingHandCursor)
        self.testmodbusrtu.setFixedSize(120, 30)
        layout.addWidget(self.testmodbusrtu, 2, 1)
        self.testmodbusrtu.clicked.connect(self.test_ModbusRTU)

        self.testmodbustcp = QPushButton("Test Mdbus TCP")
        self.testmodbustcp.setStyleSheet(
                    """
            QPushButton {
                background-color: white;
                border: None;
                border-radius: 15px; 
                padding: 8px 16px;
                font-size: 12px;
            }

            QPushButton:hover {
                background-color: #A6F1F4; 
                border-color: grey; 
            }
            """
        )
        self.testmodbustcp.setCursor(Qt.CursorShape.PointingHandCursor)
        self.testmodbustcp.setFixedSize(120, 30)
        layout.addWidget(self.testmodbustcp, 3, 0)
        self.testmodbustcp.clicked.connect(self.test_ModbusTCP)

        self.testdi = QPushButton("Test DI")
        self.testdi.setStyleSheet(
                    """
            QPushButton {
                background-color: white;
                border: None;
                border-radius: 15px; 
                padding: 8px 16px;
                font-size: 12px;
            }

            QPushButton:hover {
                background-color: #A6F1F4; 
                border-color: grey; 
            }
            """
        )
        self.testdi.setCursor(Qt.CursorShape.PointingHandCursor)
        self.testdi.setFixedSize(120, 30)
        layout.addWidget(self.testdi, 3, 1)
        self.testdi.clicked.connect(self.test_DI)

        self.testai = QPushButton("Test AI")
        self.testai.setStyleSheet(
                    """
            QPushButton {
                background-color: white;
                border: None;
                border-radius: 15px; 
                padding: 8px 16px;
                font-size: 12px;
            }

            QPushButton:hover {
                background-color: #A6F1F4; 
                border-color: grey; 
            }
            """
        )
        self.testai.setCursor(Qt.CursorShape.PointingHandCursor)
        self.testai.setFixedSize(120, 30)
        layout.addWidget(self.testai, 4, 0)
        self.testai.clicked.connect(self.test_AI)

        self.testall = QPushButton("Test All")
        self.testall.setStyleSheet(
                    """
            QPushButton {
                background-color: white;
                border: None;
                border-radius: 15px; 
                padding: 8px 16px;
                font-size: 12px;
            }

            QPushButton:hover {
                background-color: #A6F1F4; 
                border-color: grey; 
            }
            """
        )
        self.testall.setCursor(Qt.CursorShape.PointingHandCursor)
        self.testall.setFixedSize(120, 30)
        layout.addWidget(self.testall, 4, 1)
        self.testall.clicked.connect(self.test_All)

        exittestmode = QPushButton("Exit Test Mode")
        exittestmode.setStyleSheet(
                    """
            QPushButton {
                background-color: white;
                border: None;
                border-radius: 15px; 
                padding: 8px 16px;
                font-size: 12px;
            }

            QPushButton:hover {
                background-color: #A6F1F4; 
                border-color: grey; 
            }
            """
        )
        exittestmode.setCursor(Qt.CursorShape.PointingHandCursor)
        exittestmode.setFixedSize(120, 30)
        layout.addWidget(exittestmode, 5, 0, 1, 0, alignment=Qt.AlignmentFlag.AlignCenter)
        exittestmode.clicked.connect(self.exit_Test_mode)

        self.setLayout(layout)

    def test_RTC(self):
        global currentState
        self.statusbar.clearMessage()
        self.movie_label.setVisible(True)
        self.serial_thread.send_data("1" + "\n")
        currentState = STATE.TESTRTC.value

    def test_GSM(self):
        global currentState
        self.statusbar.clearMessage()
        self.movie_label.setVisible(True)
        self.serial_thread.send_data("2" + "\n")
        currentState = STATE.TESTGSM.value

    def test_WiFi(self):
        global currentState
        self.statusbar.clearMessage()
        self.movie_label.setVisible(True)
        self.serial_thread.send_data("3" + "\n")
        currentState = STATE.TESTWIFI.value

    def test_Ethernet(self):
        global currentState
        self.statusbar.clearMessage()
        self.movie_label.setVisible(True)
        self.serial_thread.send_data("4" + "\n")
        currentState = STATE.TESTETH.value

    def test_SDCard(self):
        global currentState
        self.statusbar.clearMessage()
        self.movie_label.setVisible(True)
        self.serial_thread.send_data("5" + "\n")
        currentState = STATE.TESTSD.value

    def test_ModbusRTU(self):
        global currentState
        self.statusbar.clearMessage()
        self.movie_label.setVisible(True)
        self.serial_thread.send_data("6" + "\n")
        currentState = STATE.TESTMODRTU.value

    def test_ModbusTCP(self):
        global currentState
        self.statusbar.clearMessage()
        self.movie_label.setVisible(True)
        self.serial_thread.send_data("7" + "\n")
        currentState = STATE.TESTMODTCP.value

    def test_DI(self):
        pass

    def test_AI(self):
        pass

    def test_All(self):
        global currentState
        currentState = STATE.TESTALL.value
        self.statusbar.clearMessage()
        self.movie_testAll.setVisible(True)
        self.serial_thread.send_data("8" + "\n")

    def exit_Test_mode(self):
        global currentState
        currentState = STATE.CONNECTED.value
        self.movie_label.setVisible(False)
        self.statusbar.clearMessage()
        self.serial_thread.send_data("9" + "\n")


class CalibrateAIWindow(QWidget):
    """A widget to calibrate the AI module"""
    def  __init__(self, statusbar, window_icon, serial_thread, parent=None):
        super().__init__(parent)
        self.serial_thread = serial_thread
        self.window_icon = window_icon
        self.statusbar = statusbar

        self.setWindowTitle('Calibrate AI')

        cali_layout = QGridLayout()

        self.channel = QLabel('Channel: ')
        self.channel_number = QComboBox()
        self.channel_number.setStyleSheet("QComboBox {background-color: white; border: 2px solid gray; border-radius: 10px; padding: 0 8px; }")
        self.channel_number.setPlaceholderText('Select Channel')
        self.channel_number.addItems(['0', '1', '2', '3'])
        self.channel_number.setCursor(Qt.CursorShape.PointingHandCursor)

        cali_layout.addWidget(self.channel, 0, 0)
        cali_layout.addWidget(self.channel_number, 0, 1)

        self.scale = QLabel("Scale Value: ")
        self.scale_value = QLineEdit()
        self.scale_value.setStyleSheet("QLineEdit {background-color: white; border: 2px solid gray; border-radius: 10px; padding: 0 8px; }")
        self.scale_value.setPlaceholderText("Enter Scale Value")

        cali_layout.addWidget(self.scale, 1, 0)
        cali_layout.addWidget(self.scale_value, 1,  1)

        self.offset = QLabel("Offset: ")
        self.offset_value = QLineEdit()
        self.offset_value.setStyleSheet("QLineEdit {background-color: white; border: 2px solid gray; border-radius: 10px; padding: 0 8px; }")
        self.offset_value.setPlaceholderText("Enter Offset Value")

        cali_layout.addWidget(self.offset, 2, 0)
        cali_layout.addWidget(self.offset_value, 2,  1)

        self.calibrate_button = QPushButton("Calibrate")
        self.calibrate_button.clicked.connect(self.on_calibrate_pressed)
        self.calibrate_button.setStyleSheet(
                    """
            QPushButton {
                background-color: white;
                border: None;
                border-radius: 15px; 
                padding: 8px 16px;
                font-size: 12px;
            }

            QPushButton:hover {
                background-color: #A6F1F4; 
                border-color: grey; 
            }
            """
        )
        self.calibrate_button.setFixedSize(80, 30)
        self.calibrate_button.setCursor(Qt.CursorShape.PointingHandCursor)

        cali_layout.addWidget(self.calibrate_button, 3, 0)

        self.exit_button = QPushButton("Exit Calibration Mode")
        self.exit_button.clicked.connect(self.exit_from_Calibration)
        self.exit_button.setStyleSheet(
                    """
            QPushButton {
                background-color: white;
                border: None;
                border-radius: 15px; 
                padding: 8px 16px;
                font-size: 12px;
            }

            QPushButton:hover {
                background-color: #A6F1F4; 
                border-color: grey; 
            }
            """
        )
        self.exit_button.setFixedSize(150, 30)
        self.exit_button.setCursor(Qt.CursorShape.PointingHandCursor)

        cali_layout.addWidget(self.exit_button, 3, 2)

        self.setLayout(cali_layout)

    def  on_calibrate_pressed(self):
        self.channel_no = self.channel_number.currentText()
        self.scale_val = self.scale_value.text()
        self.offset_val = self.offset_value.text()

        if not self.channel_no:
            msg_box = HandPointerMessageBox()
            msg_box.setWindowTitle("Warning")
            msg_box.setWindowIcon(self.window_icon)
            msg_box.setText("Incorrect Channel. Please Select the Channel."   )
            msg_box.setIcon(QMessageBox.Icon.Warning)
            msg_box.exec()
            return
        
        if not self.scale_val:
            msg_box = HandPointerMessageBox()
            msg_box.setWindowTitle("Warning")
            msg_box.setWindowIcon(self.window_icon)
            msg_box.setText("Incorrect Scale value. Please Enter the Scale Value ."   )
            msg_box.setIcon(QMessageBox.Icon.Warning)
            msg_box.exec()
            return
        
        if not self.offset_val:
            msg_box = HandPointerMessageBox()
            msg_box.setWindowTitle("Warning")
            msg_box.setWindowIcon(self.window_icon)
            msg_box.setText("Incorrect Offset Value. Please Enter the Offset Value."   )
            msg_box.setIcon(QMessageBox.Icon.Warning)
            msg_box.exec()
            return
        
        threading.Thread(target=self._on_calibrate_pressed).start()
    
    def _on_calibrate_pressed(self):
        self.serial_thread.send_data(self.channel_no + "\n")
        time.sleep(2)

        self.serial_thread.send_data(str(int(float(self.scale_val)*100)) + "\n")
        time.sleep(4)

        self.serial_thread.send_data(str(int(float(self.offset_val)*100)) + "\n")

        time.sleep(3)
        self.statusbar.showMessage("Calibration AI Done!")

    def exit_from_Calibration(self):
        global currentState
        currentState = STATE.CONNECTED.value
        self.serial_thread.send_data("9" + "\n")


class TerminalWindow(QWidget):
    try:
        def  __init__(self, data=None):
            super().__init__()
            self.data = data

            layout = QGridLayout()

            serial = QLabel("Serial Monitor: ")
            layout.addWidget(serial, 0, 0)

            self.serial_text = QTextEdit(readOnly = True)
            self.serial_text.setStyleSheet("QTextEdit {background-color: white; border: 2px solid gray; border-radius: 10px; padding: 0 8px; }")
            layout.addWidget(self.serial_text, 1, 0, 1, 2)

            self.setLayout(layout)

    except RuntimeError as e:
        print('Caught error: ', e)


if  __name__ == "__main__":
    app = QApplication(sys.argv)
    window = LoginWindow()
    window.setStyleSheet("background-color: #add8e6;")
    window.show()
    sys.exit(app.exec())